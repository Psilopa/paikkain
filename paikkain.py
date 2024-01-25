# Current implementation depends on dictionaries keeping their order, which is true in Python3 

# TODO: SHOULD CHECK THAT ALL CELLS IN FIRST 2(3) LINES HAVE SENSIBLE VALUES:
# (UNIQUE for 1, regular vocab for 2)
# TODO: Add support for locality text field manipulation even when there are existing coordinate fields
# TODO:  add support for checking Finnish bioregion names
# TODO: ADD output format replace_if_previously_empt
class WriteRow(Exception): pass

# Call example:     jkgeoref setup.ini inputfile.xlsx
from pathlib import Path
import sys,  shutil,  atexit,  configparser,  datetime,  time,  argparse,  os
import jksheet
from jkerror import jkError
from jktest import known_test_types 
from jktools import loadtime,  joinstr,  my2str
from openpyxl.styles import PatternFill
import logging
progname = 'paikkain'

def createlogger(fn):
    logger = logging.getLogger(progname)
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(fn)
    ch = logging.StreamHandler()
    formatter = logging.Formatter('%(message)s [%(levelname)s]')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    logger.addHandler(fh)
    logger.addHandler(ch)   
    return logger

starttime = time.time()

def onexit(): 
    log.info("Done")
    if ( ('outdata' in dir()) and  outdata ) : outdata.close()
    if ( ('geodata' in dir()) and  geodata ) : geodata.close()
    if ( ('indata' in dir()) and  indata ) : indata.close()
    endtime = time.time()
    log.info(f"Time spent: %-.2f s" % (endtime - starttime) )
#    input("Hit enter to end process: ")

atexit.register(onexit)

def dict_has_content_in(tdict,  searchkeys):
    for dc in searchkeys:
        if tdict[dc]: return True 
        else: return False
    
def row_has_date(row):
    for n in row:
        try:  
            loadtime( n ); return True # Return true is conversion to date worked
        except (ValueError,  KeyError):  pass            
    return False    
    
# Select one of several matching geodata lines
# Logic: 1. check that only date elements differ 2. prefer match with time elements.
#def match_selector(geodata,  matchrows,  rowdict,  datecolumns=[]):
#    datecolumns = ['MYGathering[0][MYDateBegin]','MYGathering[0][MYDateEnd]' ]
#    datecolumns = [x.lower() for x in datecolumns]
#    # Check that userdata has at least one date: otherwise cannot select
#    userhasdate = dict_has_content_in(rowdict,  datecolumns) 
##    if len(matchrows) > 2: print(f"userhasdate {userhasdate}")
#    no_date_rules_row = None
#    for rowi in matchrows:
#        hasdate = row_has_date(geodata.get_row_values(rowi))
#        if not hasdate: #
#            # Several matching rules with no date information: cannot sort these out, return full list as-is
#            if no_date_rules_row is not None: return matchrows
#            else: no_date_rules_row = rowi
##    if len(matchrows) > 2: print(f"no_date_rules_row {no_date_rules_row}")
#    if not userhasdate: 
#        if ( no_date_rules_row is not None): return [no_date_rules_row]
#        else: return matchrows
#    else: # User provided date: pick matching line with validity timing data
#        matchrows.remove(no_date_rules_row)
#        return matchrows

def create_output_name(infn, addition):
    infn = Path(infn)
    return infn.with_suffix(f".{addition}" + infn.suffix)

#  ------------------ main script
# Read command line
ap =argparse.ArgumentParser(description='Georeferense Excel files with geodata information')
ap.add_argument('conffn', metavar='conffn', nargs=1, help='configuration file name')
ap.add_argument('input_files', metavar='input_files', nargs='+', help='input data file(s)')
args = ap.parse_args()

executedir = Path(sys.argv[0]).parent
log = createlogger( executedir / Path(progname + ".log") )
log.info(f"Starting {progname} on {datetime.datetime.now()}")

input_files  = [Path(x) for x in args.input_files]

# Read parameters from config file
conffn = args.conffn[0]
log.info( f"Reading configuration file {conffn}"  )
config = configparser.ConfigParser(interpolation=configparser.ExtendedInterpolation() )
config.read(conffn, 'UTF-8')

insheetname = config.get('inputfiles', 'datasheetname', fallback=None)
input_first_data_line = config['inputfiles'].getint('first_data_line')
keep_original_data_marker = config['inputfiles']['keep_original_data_marker']
keep_original_data_marker = keep_original_data_marker.lower()

gdsheetname = config.get('geodatafile', 'sheetname', fallback=None)
gdfn = Path(config['geodatafile']['filename'])


cmd_replace = config['geodatafile']['cmd_replace']
cmd_append = config['geodatafile']['cmd_append']
cmd_nothing = config['geodatafile']['cmd_nothing']
outputops = [cmd_replace, cmd_append, cmd_nothing]
activeops = [cmd_replace, cmd_append]
pnote = config.get('outputfiles', 'transcribernote', fallback = "")
output_marker = config['outputfiles']['filename_add']
if config['outputfiles'].getboolean('add_date_to_note'):
    pnote = pnote + " (%s)" % datetime.date.today()
outputformat = config['outputfiles']['output_format']
outputformat = outputformat.lower()
if outputformat not in ['csv', 'xlsx',  'fast-xlsx']: 
    log.critical(f"Unknown output format: {outputformat.upper()}"); sys.exit()
    
log.info(f"Output format: {outputformat.upper()}")

if pnote: pnotecolname = config['outputfiles']['transcribernotefield']
itemsep = config['outputfiles']['data_append_connector'] + " "
replacefillcolor = config['outputfiles']['replace_fillcolor']
appendfillcolor = config['outputfiles']['append_fillcolor']
new_field_insert_point = config['outputfiles'].getint('new_column_insertion_position')

original_geodata_header = config['outputfiles']['original_geodata_to_column_header']
append_original_geodata_to_column = config.get('outputfiles', 'append_original_geodata_to_column', fallback=None)

skip_if_content_columnnames = []
for n in range(1, 10): 
    colname = 'skip_if_nonempty%i' % n
    skipn = config['inputfiles'].get(colname)
    if skipn: skip_if_content_columnnames.append(skipn.lower())

# Colour objects for XLS cell background setting
replaceFill = PatternFill(start_color=replacefillcolor, end_color= replacefillcolor, fill_type='solid')
appendFill = PatternFill(start_color=appendfillcolor, end_color= appendfillcolor, fill_type='solid')

outdata = None
geodata = None

# Read geodata file
try: 
    log.info(f"Loading geodata from file {gdfn}")
    geodata = jksheet.GeoData.fromfile(Path(gdfn), gdsheetname)     
except (FileNotFoundError,  jkError) as err: 
    log.critical(f"{err} Exiting")
    sys.exit()
log.debug("Parsing rules from geodata file headers")
rules = geodata.parse_rules(known_test_types) # Parse row matching rules from GeoData file header rows
#log.debug(f"Found the following test rules:")
for rule in rules: log.info(f"Rule for column {rule.colname}, rule type '{rule.type}'")

for infn in input_files:   
    # Read user data file
    log.info(f"\n\nProcessing file {infn}") 
    try:
        # Set up output file by copying in the input file
        outfn = create_output_name(infn, output_marker)        
        if outfn.exists(): 
            log.critical(f"File {outfn} exists. Will not overwrite. Exiting."); sys.exit()            
        shutil.copyfile(infn, outfn) # Make a copy of the original file, operate on it
        if outputformat == 'csv':
            outdata = jksheet.CSVOut.fromfile(outfn, insheetname)
            origfn = outfn
            outfn = outfn.with_suffix("out.csv") 
            if outfn.exists(): 
                log.critical(f"File {outfn} exists. Will not overwrite. Exiting."); sys.exit()            
            outdata.outputopen(outfn)
        # THIS VERSION USE IN-PLACE EXCEL EDITING: KEEPS FORMATTING, BUT IS VERY SLOW
        elif outputformat == 'xlsx': 
            outdata = jksheet.InplaceOut.fromfile(outfn, insheetname)
        elif outputformat == 'fast-xlsx': 
            outdata = jksheet.fastXLSXOut.fromfile(outfn, insheetname)
            origfn = outfn
            outfn = outfn.with_suffix(".out.xlsx") 
            if outfn.exists(): 
                log.critical(f"File {outfn} exists. Will not overwrite. Exiting."); sys.exit()            
            outdata.outputopen(outfn)
            
        indata = outdata        

        # Add fields to output table as needed on the basis of geodata file header
        for colname in geodata.output_column_names(activeops):
            if not outdata.hascolumn(colname):
                log.info(f"adding column {colname} to output table")
                outdata.addcolumn( colname, new_field_insert_point )
        if pnotecolname and pnote:
            if not outdata.hascolumn(pnotecolname):
                log.info(f"adding column {pnotecolname} to output table")
                outdata.addcolumn( pnotecolname, new_field_insert_point )
        if append_original_geodata_to_column:
            if not outdata.hascolumn(append_original_geodata_to_column):
                log.info(f"adding column {append_original_geodata_to_column} to output table")
                outdata.addcolumn(append_original_geodata_to_column , new_field_insert_point )                
            original_geodata_col = outdata.colnumber(append_original_geodata_to_column) # Note; this must be last insertion, otherwise we need to update this
            
        # Copy header lines 
        outdata.copyheaders(input_first_data_line) # Copy header lines to possible alternative output files

        # Step through input file and process line by line
        for row in range( input_first_data_line, indata. nrows +1 ): 
            if (row % 10) == 0: log.info(f"Processing row {row}") 
            rowdict = indata.get_row_as_dict(row) 
            outdict = indata.get_row_as_dict(row) 
            edited = { k: False for k in  rowdict.keys() }
            try: 
                for skipname in skip_if_content_columnnames: 
                    if not outdata.isempty_by_colname(row,  skipname): 
                        raise WriteRow
                matchrows = geodata.find_matches(rowdict,  rules)
    #            if len(matchrows) > 1: matchrows = match_selector(geodata,  matchrows,  rowdict)
                nmatch = len(matchrows)
                if nmatch == 0:  
                    raise WriteRow
                if nmatch > 1:  
                    log.debug(f"Found multiple matches for inputrow {row}: {matchrows}. Check geodata source file. Skipping row")
                    raise WriteRow
                # OK, so we have exactly one match
                try:
                    originaldata = []
                    mrow = matchrows[0] # index of matching row
                    match = geodata.get_row_as_dict( mrow )
                    for colname,val in match.items():  # Iterate over columns in match item
                        # If column name is not in outdata, it is not an active output field name and can be ignored
                        if colname.lower() not in outdata.lowercolnames: continue                         
                        col = indata.colnumber(colname)
                        oval = indata.getvalue(row, col)  # Value in input data at this position
                        if my2str(val).strip().lower() == keep_original_data_marker: continue
                        # Copy original data to a field in the output file (not copying the output cell data into itself
                        if append_original_geodata_to_column and (col != original_geodata_col):  
                            if oval: originaldata.append( str(indata.getvalue(row, col) ) )
                        oper = geodata.get_output_action_for_column(colname, outputops) 
                        if oper not in outputops:
                            continue # Skip column with actions that are not output operations
                        elif oper == cmd_replace:
                            outdict[colname] = val
                            edited[colname] = True
                            outdata.setbackground(row, col, replaceFill)
                        elif oper == cmd_append and val: # Append non-empty values only
                            outdict[colname] = joinstr( outdict[colname] or "",  val ,  itemsep ) 
                            edited[colname] = True
                            outdata.setbackground(row, col, appendFill)
                    if append_original_geodata_to_column: # Append old data to designated cell
                            origstr = f"\n{original_geodata_header} {itemsep.join(originaldata)}" 
                            cn = append_original_geodata_to_column.lower()
                            outdict[cn] = joinstr(outdict[cn] or "",  origstr ,  "") 
                            edited[cn] = True
                            outdata.setbackground(row, original_geodata_col, appendFill)
                    # Add note by the program, if available
                    if pnotecolname and pnote:
                        cn = pnotecolname.lower()
                        outdict[cn] = joinstr(outdict[cn] or "",  pnote ,  itemsep) 
                        edited[cn] = True
                        outdata.setbackground(row, outdata.colnumber(cn), appendFill)
                    raise WriteRow
                except (jkError) as msg:
                    log.critical(msg)
                    sys.exit()
            except WriteRow: 
                outdata.writerow(row,  outdict,  edited)
        log.info(f"Saving output file {outfn}") 
        outdata.save()
        outdata.close()
        if outputformat in ['csv',  'fast-xlsx'] and origfn.exists(): os.remove(origfn) 
    except (jkError,  FileNotFoundError) as msg:
        log.critical(msg)
        sys.exit() 

# Ask for any input before closing window
# Now handled to a OS script wrapper (.bat on Windows)