# A wrapper around spreadsheet-like files (or other similar objects)
#import pathlib import Path
import openpyxl     
from openpyxl.cell import WriteOnlyCell
import jktest,  jktools
from jkerror import jkError
import csv, logging, abc

progname = 'paikkain'
log = logging.getLogger(progname)

editedFill = openpyxl.styles.PatternFill(start_color="fa867e", end_color= "fa867e", fill_type='solid')

    
class jkXLSfilesheet(): 
    """Note that all indexing in this class is 'Excel-style', ie. first row = 1, first column = 1"""
    def __init__(self):
#        self.colnames = None # A row of column titles
        self.origcolnames = None # A row of column titles as they are in the file
        self.lowercolnames = None # A row of column titles, lowercase
        self.rulerow = None # A row of column types/rules
        self.ws = None #  working sheet 
        self.wb =None # workbook
        self.fn = None # filename
        self.ncols = 0
        self.nrows = 0
        self.valuegrid = None # In-memory copy of cell values only for speed


    def open(self,  fn,  sheetname=None):
        """Uses first sheet is name is not given"""
        self.fn = fn
        self.valuegrid = None
        self.wb = openpyxl.load_workbook( fn,  data_only =True)  
        if not sheetname: sheetname = self.wb.sheetnames[0]
        if not sheetname in self.wb.sheetnames: 
            raise jkError(f"Unknown sheet name '{sheetname}' in file {fn}")
        self.ws = self.wb[sheetname]
        self.ncols = self.ws.max_column
        #print(f"NUMBER OF COLUMNS IN {self.fn} = {self.ncols}")
        self.nrows = self.ws.max_row
        if self.nrows < 3:  raise jkError(f"File {fn} too short, needs at least two header line and one data line")
        # Look for empty columns
        emptycols = []
        for col in range(1, self.ncols+1):
            if self.column_is_empty(col): emptycols.append(col)
        for col in emptycols[::-1] :
            log.debug("Deleting empty column #%s" % col )
            self.column_delete(col)  # Delete columns starting from the end        
        # Local memory copy of values for speed: create line 0 to match row indexing with self.ws (which indexes starting from 1)
        self._valuegrid_from_data() # Copy data to a grid for read access speed!
        self._updateheadervars()
        self._indexcolumnnames()

    def _valuegrid_from_data(self):
        self.valuegrid =  [[None]*self.nrows, ] 
        for row in self.ws.values: self.valuegrid.append([x for x in row])

    def _indexcolumnnames(self):
        self.n2c = {self.lowercolnames[n]: n  for n in range(1,self.ncols)}
        return self.n2c
        
    def close(self):
        if hasattr(self, 'wb') and self.wb: self.wb.close()
        if hasattr(self, 'nwb') and self.wb: self.nwb.close()
        self.wb = None
        
    def _updateheadervars(self):
        self.origcolnames = self.get_row_values( 1 ) 
        self.origcolnames = [ x.strip() for x in self.origcolnames ] 
        for i in range(self.ncols):
            if self.origcolnames[i] == '':
                for i in range(self.ncols): print(i+1, self.origcolnames[i]) # Output column name list before exiting with an Exception
                raise jkError(f"Empty values on first row not allowed. Every column should have a name. Check column {i+1}")
        self.lowercolnames = [ x.lower() for x in self.origcolnames ]
        self.rulerow =  [ str(x) for x in self.get_row_values( 2 ) ]
        self.rulerow = [ x.strip() for x in self.rulerow ]

    # Column operations
    
    def isempty_by_colname(self, row,  colname): 
        cn = self.colnumber(colname)        
        return self.isempty(row,  cn)
            
    def hascolumn(self,  colname,  casesensitive=False):
        if not casesensitive: 
             return ( colname.lower() in self.lowercolnames )
        else: raise jkError("Case sensitive column name lookup not implemented")
        
    def column_get_values(self, col): 
        return next( self.ws.iter_cols(min_col=col, max_col=col,  values_only =True)  ) # Return first column from the iter_cols iterator

    def column_is_empty(self, col): 
        for val in self.column_get_values(col):
            if val is not None: return False
        return True

    def column_delete(self, col):  
        self.ws.delete_cols(col, 1)
        if self.valuegrid: 
           for row in self.valuegrid: del row[col] 
        self.ncols -= 1
                    
    def colnumber(self, name):
        """Number in column list corresponding to column name 'name'"""
        return self.n2c[name.lower()]
#    def originalcolname(self, n): return  self.origcolnames[n-1] # origcolnames is 0-indexed!
#    def lowercolname(self, n): return  self.lowercolnames[n-1] # lowercolnames is 0-indexed!

    # Row operations
    def get_row_values(self, row): return [x  or ''  for x in self.valuegrid[row] ]
    
    def get_row_as_dict(self, rowi): 
        """Return a row converted to a dictionary with column header (line 1): cell value pairs. Excel-style indexing, ie first row = 1"""
        return { k: v for (k, v) in zip(self.lowercolnames, self.valuegrid[rowi] ) }        

    # Cell operations
    def getvalue(self, row, col): # Todo: rename to getCellValue
        return self.valuegrid[row][col]
        
    def getvalue_by_colname(self, row, colname): 
        """Returns first occurence of colname, if it repeats!"""
        return self.getvalue(row, self.colnumber(colname) )
        
    def isempty(self, row,  col): 
        if (self.getvalue(row, col) == None) or \
             (self.getvalue(row, col) == '') or \
             (self.getvalue(row, col) == []): return True
        else: return False

    # Cell operations for multivalue cells
    # TODO: convert all getvalue-type calls to return a list of values
    def getvaluelist(self, row, col): # Todo: rename to getCellValueList
        return self.valuegrid[row][col]
        
    def getvaluelist_by_colname(self, row, colname): 
        """Returns cell values as a list (for multi-value cells)

        Returns first occurence of colname, if it repeats!"""
        return self.getvalue(row, self.colnumber(colname) )        

##class jkXLSreadonly(jkXLSfilesheet): 
##    def __init__(self,  *args, **kwargs):        
##        super(jkXLSfilesheet, self).__init__(*args, **kwargs)        
##        
##    def open(self,  fn,  sheetname): 
##        self.fn = fn
##        self.wb = openpyxl.load_workbook( fn,  data_only =True)
##        if not sheetname in self.wb.sheetnames: 
##            raise jkError(f"Unknown sheet name '{sheetname}' in file {fn}")
##        self.ws = self.wb[sheetname]
###        self.ws.reset_dimensions()
##        self.ncols = self.ws.max_column
##        self.nrows = self.ws.max_row
##        if self.nrows < 3: 
##            raise jkError(f"File {fn} too short, needs at least two header line and one data line")
##        # Local memory copy of values for speed: create line 0 to match row indexing with self.ws (which indexes starting from 1)
##        self._valuegrid_from_data() # Copy data to a grid for read access speed!
##        self._updateheadervars()
##        self._indexcolumnnames()        
        
        
class GeoData(jkXLSfilesheet):
    first_data_row = 4 # Excel indexing, starts a 1!

    def __init__(self,  *args, **kwargs):        
        super(jkXLSfilesheet, self).__init__(*args, **kwargs)        
        
    @classmethod
    def fromfile(GeoData, fn,  sheetname):
        x = GeoData()
        x.open(fn,  sheetname)
        return x

    def parse_rules(self,rulenames):
        """Find columns with rules and store them. rulenames = a list of allowed rule names."""
        rules = []
        for i in range(0,  self.ncols): #            
            if self.lowercolnames[i] is None: continue  # Skip cols with nothing in title row            
            inrule = self.rulerow[i].strip().lower()
            if inrule not in rulenames: continue
            rule = jktest.singlerule( i, self.origcolnames[i], inrule)
            rules.append(rule)                
        return tuple(rules)
        
    def find_matches(self, usedatadict,  rules): 
        """Recturns a list of indices to matching rows """
        matches = []         
        # Standardize value: no double spaces, no .;:
        udd = {k: jktools.locstrip(v) for (k, v ) in usedatadict.items()}
        # Test each known georef row and look for perfect matches for all tests.         
        try:
            for row in range(self.first_data_row,self.nrows+1):
                testsuccesses = 0 
                matchall = True
                for rule in rules: 
                    test_against_value = self.getvalue( row,  rule.col ) # Get geodata value to test user data against
                    testresultcode = rule.match(udd,  test_against_value)
                    testsuccesses += testresultcode
                    if ( testresultcode == 2 ): # Failure to match
                        matchall = False 
                        break   
                if matchall and (testsuccesses > 0) :  # If testsuccess == 0, all tests defaulted to success because there was no data to test                
#                    matches.append(self.get_row_as_dict(row) )
                    matches.append( row )
        except ValueError: pass
        return matches
        
    def get_output_action_for_column(self, column_name,  acceptedtypes):
        # Find in which column it occurs in with an accepted output type
        column_name = column_name.lower()
        for n in range(len(self.lowercolnames)):
            outaction = self.rulerow[n]
            if (self.lowercolnames[n] == column_name) and (outaction in acceptedtypes): 
                return outaction
        return None # If no suitable column was found
    
    def output_column_names(self,  actions=[]):
        # Look throiugh
        names = []
        for i in range(len(self.rulerow)):
            if self.rulerow[i] in actions: names.append(self.origcolnames[i])
        return names
        
                
# BASIC INPUT FILE CLASS
class InputData(jkXLSfilesheet, abc.ABC): # Currently a thin wrapper around Excel sheets, but may change
    def __init__(self,  *args, **kwargs):
        super(jkXLSfilesheet, self).__init__(*args, **kwargs)        
    @classmethod
    def fromfile(InputData, fn,  sheetname):
        x = InputData()
        x.open(fn,  sheetname)
        return x


# BASIC OUTPUT FILE CLASS
class OutData(jkXLSfilesheet, abc.ABC):
    def __init__(self,  *args, **kwargs):
        super(jkXLSfilesheet, self).__init__(*args, **kwargs)        

    def save(self):
        if self.wb: self.wb.save( self.fn )
        else: raise jkError(f"Save failed: no file data available to save")

    # Column operations
    def addcolumn(self, name, new_column_insert_index=1):
        name = name.strip()
        self.ws.insert_cols(new_column_insert_index)
        self.setvalue(1, new_column_insert_index -1,  name)
        self.ncols = self.ws.max_column
        self._valuegrid_from_data()
        self._updateheadervars()
        self._indexcolumnnames()
        
    # Cell operations
    def setvalue(self, row, col,  val):
        self.ws[row][col].value = val
        self.valuegrid[row][col] = val
        
    def setmultivalue(self, row, col, val, sep=";"):
        # Flattens value for the Excel sheet representation, if value is a list of string
        self.ws[row][col].value = sep.join(val) or ""
        self.valuegrid[row][col] = val
        
    def replacevalue(self, row, col, val):
        self.setnumberformat(row, col, '@')
        self.setvalue(row, col,  val)
    
    def appendvalue(self, row, col, val,  sep=", "): 
        self.setnumberformat(row, col, '@')
        if not val: return  # No nothing for None or "" as input
        if self.isempty(row, col): self.setvalue(row, col,  val)
        else: 
            oldval = str( self.getvalue(row, col) )
            newval = sep.join( [oldval , val]  ) 
            self.setvalue(row, col,  newval)

    # Override thesein subclasses, if operation is meaningful
    # TODO: make these actually virtual
    @abc.abstractmethod
    def copyheaders(self,  nrows): pass 
    @abc.abstractmethod
    def setnumberformat(self, row, col, format): pass
    @abc.abstractmethod
    def setbackground(self, row, col, fillcolor): pass
    @abc.abstractmethod
    def writerow(self, row, edited): pass

                    
class CSVOut(OutData):
    def __init__(self,  *args, **kwargs):
        super(OutData, self).__init__(*args, **kwargs) 
    @classmethod
    def fromfile(OutData, fn,  sheetname):
        x = OutData()
        x.open(fn,  sheetname)
        return x
    def outputopen(self, fn):
        self.csvfile = open(fn, 'w')
        self.wr = csv.writer(self.csvfile,  delimiter=";",  quotechar='"', quoting=csv.QUOTE_MINIMAL)
    def copyheaders(self,  nrows): 
        for i in range(1, nrows):
            self.wr.writerow( self.get_row_values(i) )
    def writerow(self, row,  rowdict, edited): 
        strvals =  [ (x or '') for x in rowdict.values() ]   
        strvals = [str(x) for x in strvals]
        self.wr.writerow( strvals )
    # TODO: make these actually virtual
    @abc.abstractmethod
    def copyheaders(self,  nrows): pass 
    def setnumberformat(self, row, col, format): pass
    def setbackground(self, row, col, fillcolor): pass
    @abc.abstractmethod
    def writerow(self, row, edited): pass


class fastXLSXOut(OutData):
    def __init__(self,  *args, **kwargs):
        super(OutData, self).__init__(*args, **kwargs) 
    @classmethod
    def fromfile(fastXLSXOut, fn,  sheetname):
        x = fastXLSXOut()
        x.open(fn,  sheetname)
        return x
    def outputopen(self, fn):
        self.nfn = fn
        self.nwb = openpyxl.Workbook(write_only = True)
        self.nws = self.nwb.create_sheet()
    def copyheaders(self,  nrows): 
        for i in range(1, nrows):
            vals = ( str(x) for x in self.get_row_values(i)  )
            self.nws.append(vals)
#    def _writerow(self, row): 
#        self.nws.append(row)
    def writerow(self, row,  rowdict, edited): 
        cellrow = []
        for k in rowdict.keys():            
            cell = WriteOnlyCell(self.nws, value= rowdict[k])
            cell.number_format = '@' # TEXT
            if edited[k]:  cell.fill = editedFill               
            cellrow.append(cell)
        self.nws.append(cellrow)    
    def close(self):
        self.nwb.save(self.nfn)
        super(OutData, self).close()
    def setnumberformat(self, row, col, format): pass
    def setbackground(self, row, col, fillcolor): pass

    
