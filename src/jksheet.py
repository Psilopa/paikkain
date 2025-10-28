# A wrapper around spreadsheet-like files (or other similar objects)
import openpyxl     
import jktest,  jktools
from jkerror import jkError
import csv, logging, abc
from pathlib import Path

progname = 'paikkain'
log = logging.getLogger(progname)

editedFill = openpyxl.styles.PatternFill(start_color="fa867e", end_color= "fa867e", fill_type='solid')

#Could be used to format cells depending on the operation type
op_replaced = 1
op_appended = 2
validops = [op_replaced, op_appended]

class jkExcel(abc.ABC):
    """First row, first col = 1"""
    def __init__(self,filename,first_data_line_number):     
        self.fp = Path(filename)
        self._name2col = {} 
        self._lowern2col = {} # Should use lowercase column names
        self.crow = 1 # Current row
        self.wb = self._openwb()
        self.sheet = self.wb.active # Active sheet in the workbook
        self.first_data_line = first_data_line_number
    @property
    def nrows(self): return self.sheet.max_row
    @property
    def ncols(self): return self.sheet.max_column
    @property
    def filename(self): return self.fp
    @property
    def colnames(self):
        return self._name2col.keys()
    @property
    def lowercolnames(self):
        return self._lowern2col.keys()
    @abc.abstractmethod
    def _openwb(self): 
        # Should return an open openpyXLSX workbook
        pass
    def close(self): 
        if self.wb: self.wb.save(self.fp)
    def hascolumn(self,colname, casesensitive=False): 
        if casesensitive: return colname in self.colnames
        else: return colname.lower() in self.lowercolnames
    def _update_name2column(self):
        _cotitles = tuple(( x.value for x in self.sheet[1]) )
        _ind = tuple(range(1,len(_cotitles)+1))
        self._name2col = { c:i for (c,i) in zip(_cotitles,_ind) if c}
        self._lowern2col = { c.lower():i for (c,i) in zip(_cotitles,_ind) if c}
    # Support simple iteration over lines
    def __next__(self): self.crow += 1
    def next(self): self.__next__()
    def end(self): 
        if self.crow > self.sheet.max_row: return True
        else: return False

# --------------------------- OUTPUT FILES ---------------------------
class woExcel(jkExcel):
    """First row, first col = 1"""    
    supports_fill = True
    def __init__(self,filename,first_data_line):     
        self.fill_edited= None
        super().__init__(filename,first_data_line)
    def _openwb(self): 
        # Create a new workbook
        wb = openpyxl.Workbook()
        self.sheet = wb.create_sheet()
        return wb
    # PROPERTY ATTRIBUTES
    def fill_edited_color(self, color):
        self.fill_edited = openpyxl.styles.PatternFill(start_color=color, end_color= color, fill_type='solid')
    # FILE CONTENT MODIFICATION
    def save(self): self.wb.save(str(self.fp))
    def addcolumn(self,position,header_rows=["New column"]):
        """Add column. Note: First column is 1 etc."""
        if not header_rows[0] or not header_rows[0].strip(): 
            raise ValueError(f"Missing of empty column names are not allowed.")
        if header_rows[0].lower() in self._lowern2col: 
            raise ValueError(f"Column name {header_rows[0]} does already exist")
        self.sheet.insert_cols(position) 
        for ii in range(len(header_rows)):  
            self._setcell(position,1+ii,header_rows[ii])
        self._update_name2column() 
    def _setcell(self,colno,rowno,value): # Get value
        self.sheet.cell(column=colno,row=rowno).value = value
    def _setncell(self,colname,rowno,value): # Get value by column name
        colno = self._lowern2col[colname.lower()]    
        self._setcell(colno,rowno,value)
    def iterset(self,colno,value): self._setncell(colno,self.crow,value)
    def iternset(self,colname,value): self._setncell(colname,self.crow,value)
    def itersetrow(self, dict_column_and_value,  edited):
        """Create a new row from a dctionary with column names as keys
    
    The keyword edited, of provided, must be an array of len(dict_column_and_value), with values controlling cell formatting. 0 for no formatting, 
    True values (like non-zero)  integer do currently result in cess color
        """
        cellrow = [None]*len(self.colnames)
#        print("len cellrow = ",  len(cellrow))
        for colname in dict_column_and_value:       
            pos = self._lowern2col[colname.lower()] -1
#            print(pos, colname)
            cell = openpyxl.cell.WriteOnlyCell(self.sheet, value= dict_column_and_value[colname])
            cell.number_format = '@' # TEXT
            if edited and edited[colname] and self.fill_edited: cell.fill = self.fill_edited
            cellrow[pos] = cell
        self.sheet.append(cellrow)    


# --------------------------- INPUT FILES, READ ONLY ---------------------------
class roExcel(jkExcel):
    supports_fill = False
    def __init__(self,filename,first_data_line):     
        super().__init__(filename,first_data_line)
        self._update_name2column()
    def _openwb(self): 
        wb = openpyxl.load_workbook(self.fp) #, read_only=True
        self.rows = wb.active.values # Iterator!
        return wb
#    def set_active_sheet(self,n): # Zero-based indexing
#        wb.active = n        #Should reset Next counter
#        self.sheet = self.wb.active
    def next_row(self):
        self.next()
        return next(self.rows)
    def _row2dict(self,row):
        return { k:(v or "") for (k,v) in zip(self.lowercolnames,row) }
    def next_row_as_dict(self):
        row = next(self.rows)
        self.next()
        return self._row2dict(row)
    def get_row(self,n):# Direct row access, 1-based indexing
        return [ x.value for x in self.sheet[n] ]
    def get_value(self,nrow,ncol):# Direct cell access, 1-based indexing
        return self.sheet.cell(row=nrow, column=ncol).value
    def get_row_as_dict(self,nrow):
        return self._row2dict( self.get_row(nrow) )
    
# --------------------------- RO Excel for Geodata ---------------------------

class GeoData(roExcel):
    def __init__(self,  *args, **kwargs):        
        super(roExcel, self).__init__(*args, **kwargs)
        # TODO: Hardcoded for now, move to config!
        self._row_colnames = 1
        self._row_rules = 2 # Excel indexing, 2nd row!
        self.first_data_row = 4 # Default first data row
        self._strdata = None # A tuple with data, preconverted to str for speed
        self._rulesrow = None
        self._colnamesrow = None
        
    @classmethod
    def fromfile(GeoData, fn,  sheetname, first_data_row=4):        
    # TODO: NO OPTION TO PASS SHEET NAME
        fp = Path(fn)
        if not fp.exists(): raise jkError(f"File {fp} does not exist")
        x = GeoData(fn,first_data_row)
        
        x.first_data_row = first_data_row
        x._colnamesrow = [ s or "" for s in x.get_row(x._row_colnames) ] # Replaces None with ""
        x._colnamesrow = [ s.strip() for s in x._colnamesrow ]
        x._rulesrow = [ s or "" for s in x.get_row(x._row_rules) ] # Replaces None with ""
        x._rulesrow = [ s.strip() for s in x._rulesrow ]
        x.reverse_column_names = [ s.lower() for s in x._colnamesrow[::-1] ]        
        x.reverse_rules = x._rulesrow[::-1]
#        print("collrules = ", x.reverse_column_names)
#        print("rules = ", x.reverse_rules)
        assert len(x.reverse_column_names) == len(x.reverse_rules)
        return x

    @property
    def rulesrow(self): return self._rulesrow
    
    @property
    def colnamesrow(self): return self._colnamesrow
        
    def parse_rules(self,rulenames):        
        """Find columns with rules and store them. rulenames = a list of allowed rule names."""
        self._update_name2column() # Make sure we have an index of column names
        rules = []
        lowercolnames = [x.strip().lower() for x in self.colnamesrow]
        ruleslower = [x.strip().lower() for x in self.rulesrow]
        for i in range(0,  self.ncols): #
            if not lowercolnames[i]: continue  # Skip cols with nothing in title row
            inrule = ruleslower[i]
            if inrule not in rulenames: continue
            rule = jktest.singlerule( i, self.colnamesrow[i], inrule)
            rules.append(rule)                
        return tuple(rules)

    def get_result_dict(self,nrow,acceptedtypes):
        cols = [x.lower() for x in self._colnamesrow]
        row = self.get_row(nrow)
        rules = self.rulesrow
        cols = zip(cols ,row, rules)
        retval = { k:(v or "") for (k,v,r) in cols if r and (r in acceptedtypes) }
        return retval

    def get_data_rows(self):
        if not self._strdata:
            self._strdata = list(self.sheet.values)
            for n in range(self.first_data_row): del self._strdata[0] # skip header lines
            self._strdata = tuple(self._strdata)
        return self._strdata

    def find_matches(self, datadict,  rules, ignorechars="",normalize_dict={}): 
        """Recturns a lregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsregular_subsist of indices to matching rows """
        matches = []         
        # Standardize value: no double spaces, no .;:
        normalized_data_row = {k: jktools.loc_normalize(v,ignorechars,normalize_dict) for (k, v ) in datadict.items()}
        # Test each known georef row and look for perfect matches for all tests.
        nrow = 0
        try:
            for row in self.get_data_rows():
                nrow += 1                
                testsuccesses = 0 
                matchall = True # Match all rules (rule1 AND rule2 AND ...)
                for rule in rules:
                    test_against_value = row[rule.col] # Get geodata value to test user data against
                    testresultcode = rule.match(normalized_data_row, test_against_value)
                    testsuccesses += testresultcode
                    if ( testresultcode == 2 ): # Failure to match
                        matchall = False
                        break   
                if matchall and (testsuccesses > 0) :  # If testsuccess == 0, all tests defaulted to success because there was no data to test
                    matches.append(nrow + self.first_data_line) # Store row number of matching row (correct for skipperd header lines)
        except ValueError: pass        
        return matches
        
    def get_output_action_for_column(self, column_name, acceptedtypes):
        """Find in which column it occurs in with an accepted output type.

        Arguments:
        - acceptedtypes: a list of accepted output types ('replace', 'append')
        """
        column_name = column_name.lower()
        # NOTE: THIS DEPENDS ON RULE COLUMNS BEING BEFORE OUTPUT COLUMNS
        # REVERSE BOTH LISTS
        assert column_name in self.reverse_column_names # Mismatch between rule and match column title, should not happen
        try:
            n = self.reverse_column_names.index(column_name) # Find index where column_name last occurs (first in reserved)
            outaction = self.reverse_rules[n]
            if outaction in acceptedtypes:
#                print("Found action ", outaction, "for ", column_name)
                return outaction
            else: return None
        except ValueError: # If no suitable column with a recognised output type found was found 
            return None
    
    def output_column_names(self,  actions=[]):
        names = []
        for n in range(len(self.rulesrow)):
            if self.rulesrow[n] in actions: names.append(self.colnamesrow[n])
        return names
    
