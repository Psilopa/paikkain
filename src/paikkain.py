# Call example:     jkgeoref setup.ini inputfile.xlsx
# Current implementation depends on dictionaries keeping their order, which is true in Python3 

import sys
if sys.version_info >= (3, 11):
    import tomllib
else:
    import tomli as tomllib

import atexit,  datetime,  time,  argparse
import jksheet
from jkerror import jkError
from jktest import known_test_types 
from jktools import joinstr,  my2str
from openpyxl.styles import PatternFill
from pathlib import Path
import logging
progname = 'paikkain'
version = '2.95'

starttime = time.time()
log = None # Overidden by the createlogger() call
op_replaced = 1
op_appended = 2
_SUPPRESS_FILE_CREATION_FOR_TESTING = False
first_data_line_of_geodata = 4

def createlogger(fn):
    logger = logging.getLogger(progname)
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(fn)
    ch = logging.StreamHandler()
    formatter = logging.Formatter('%(message)s [%(levelname)s]')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    logger.addHandler(fh)
    logger.addHandler(ch)   
    return logger

class WriteRow(Exception): pass

def onexit(): 
    if log: log.info("Done")
    if ( ('outdata' in dir()) and  outdata ) : outdata.close()
    if ( ('geodata' in dir()) and  geodata ) : geodata.close()
    if ( ('indata' in dir()) and  indata ) : indata.close()
    endtime = time.time()
    if log: log.info("Time spent: %-.2f s" % (endtime - starttime) )

def create_output_name(infn, addition):
    infn = Path(infn)
    return infn.with_suffix(f".{addition}" + infn.suffix)

def read_TOML_config(confname):
    """Read a configuration file in TOML. Raise a jkError on error."""    
    if not conffn.is_file(): raise jkError(f"Config file '{conffn.absolute()}' does not exist or if not readable.")    
    try:
        with conffn.open("rb") as f: config = tomllib.load(f)
    except tomllib.TOMLDecodeError as msg: raise jkError(f"Config file '{conffn.absolute()}' parsing failed: f{msg}.")
    # UGLY HARDCODED COMPONENT REPLACEMENT!
    pm = config.get("programname","")
    ver = config.get("version","")
    if not "filenames" in config["knowndatafiles"]: raise ValueError("Config file has no knowndatafiles:filenames item.")
    file = config["knowndatafiles"]["filenames"][0]    
    config["outputfiles"]["transcribernote"] = config["outputfiles"]["transcribernote"].replace("{programname}", pm)
    config["outputfiles"]["transcribernote"] = config["outputfiles"]["transcribernote"].replace("{version}", ver)
    config["outputfiles"]["transcribernote"] = config["outputfiles"]["transcribernote"].replace("{knowndatafiles:filenames}", file)    
    return config    

#  ------------------ main script
atexit.register(onexit)

# READ CONFIGURATION AND KNOWN DATA FILES
if __name__ == '__main__':
    try: 
        # Read command line
        ap =argparse.ArgumentParser(description='Georeferense Excel files with geodata information')
        ap.add_argument('conffn', metavar='conffn', nargs=1, help='configuration file name')
        ap.add_argument('input_files', metavar='input_files', nargs='+', help='input data file(s)')
        args = ap.parse_args()

        executedir = Path(sys.argv[0]).parent
        log = createlogger( executedir / Path(progname + ".log") )
        log.info(f"Starting {progname} on {datetime.datetime.now()}")
        input_files  = [Path(x) for x in args.input_files]

        # Read parameters from config file
        conffn = Path(args.conffn[0]) 
        log.info( f"Reading configuration file {conffn}"  )
        c = read_TOML_config(conffn) # CONFIG DATA 

        inc_sheetname = c['inputfiles'].get('sheetname', None)
        inc_first_data_line = c['inputfiles'].get('first_data_line', 2)
        ignorechars = c.get('ignore_in_comparison',"")
        # Regular expression replacements
        if 'replacements' in c['inputfiles']:
            regular_subs = c['inputfiles']['replacements']
        else: regular_subs = {}

        knownd_keep = c['knowndatafiles'].get('keep_original_data_marker').lower()
        knownd_sheetnames = c['knowndatafiles'].get('sheetname',None)
        knownd_filenames = [ Path(x) for x in  c['knowndatafiles'].get('filenames') ]  
        try:
            cmd_replace = c['knowndatafiles']['cmd_replace']
            cmd_append = c['knowndatafiles']['cmd_append']
            cmd_fillempty = c['knowndatafiles']['cmd_fillempty']
            cmd_nothing = c['knowndatafiles']['cmd_nothing']
        except LookupError as msg: raise jkError(f"Command names must be defined in the config file: {msg}.")
        outputops = [cmd_replace, cmd_append, cmd_fillempty, cmd_nothing]
        activeops = [cmd_replace, cmd_append, cmd_fillempty]

        pnote = c['outputfiles'].get('transcribernote', "") 
        if c['outputfiles'].get('transcribernote_appendfilenames', False):
            pnote += " "
            pnote += ", ".join((str(x) for x in knownd_filenames) )
        output_marker = c['outputfiles'].get('filename_add')
        if c['outputfiles'].get('add_date_to_note'):
            pnote = pnote + " (%s)" % datetime.date.today()
        outputformat = c['outputfiles'].get('output_format')
        outputformat = outputformat.lower()
        if outputformat not in ['csv', 'xlsx',  'fast-xlsx']: 
            log.critical(f"Unknown output format: {outputformat.upper()}"); sys.exit()
            
        log.info(f"Output format: {outputformat.upper()}")

        if pnote: pnotecolname = c['outputfiles'].get('transcribernotefield')
        itemsep = c['outputfiles'].get('data_append_connector') + " "
        new_field_insert_point = c['outputfiles'].get('new_column_insertion_position')

        original_geodata_header = c['outputfiles'].get('original_geodata_to_column_header')
        append_original_geodata_to_column = c['outputfiles'].get('append_original_geodata_to_column',None)
        skip_if_content_columnnames = c['inputfiles'].get('skip_if_nonempty',[]) # Empty list default
        skip_if_content_columnnames = [x.lower() for x in skip_if_content_columnnames]
        log.debug(f"skipping row if content if found in columns {skip_if_content_columnnames}")

        # Data file object placeholders
        outdata = None
        geodata = None

        # Read geodata files
        for knowdatafn in knownd_filenames:
            geodatalist = []
            log.info(f"Loading geodata from file {knowdatafn}")
            geodata = jksheet.GeoData.fromfile(Path(knowdatafn), 
                knownd_sheetnames, 
                first_data_line_of_geodata)     
            log.debug("Parsing rules from geodata file headers")
            rules = geodata.parse_rules(known_test_types) # Parse row matching rules from GeoData file header rows
            geodatalist.append( geodata )
            #log.debug(f"Found the following test rules:")
            for rule in rules: log.info(f"Rule for column {rule.colname}, rule type '{rule.type}'")

    except (FileNotFoundError,  jkError) as err: 
        log.critical(f"{err} Exiting.")
        sys.exit()

    # PROCESS INPUT FILES 
    for infn in input_files:   
        # Read user data file
        log.info(f"\n\nProcessing file {infn}") 
        try:
            # OPEN INPUT DATA
            indata = jksheet.roExcel(infn, inc_first_data_line)            
            # SETUP FOR OUTPUT FILE
            outfn = create_output_name(infn, output_marker)        
            if outputformat == 'fast-xlsx': 
                outfn = outfn.with_suffix(".out.xlsx") 
                outdata = jksheet.woExcel(outfn,inc_first_data_line)
                outdata.fill_edited_color("fa867e")
            else:
                log.critical(f"Unknown output format {outputformat} exists. Exiting."); sys.exit()            
#            elif outputformat == 'csv':
#                outdata = jksheet.CSVOut.fromfile(infn, inc_sheetname) # Copy existing data
#                origfn = outfn
#                outfn = outfn.with_suffix("out.csv") 
            if outfn.exists(): 
                log.critical(f"File {outfn} exists. Will not overwrite. Exiting."); sys.exit()            

            firstrow = indata.next_row()
            # Verify that 1st line is valid
            for i in range(len(firstrow)):
                if not firstrow[i]: 
                    raise ValueError(f"File {indata.filename}, column {i+1}: Empty column name (first row) not allowed.")

            # ADD COLUMNS INTO OUTPUT IF CONFIG OR KNOWN DATA CONTAIN COLS NOT IN INPUT
            for name in firstrow[::-1]:  # Grab first row. Reverse order to as insertion re-reverses them
                    outdata.addcolumn(1,  [name])                
            for colname in geodata.output_column_names(activeops)[::-1]:  # Reverse order to as insertion re-reverses them
                if not outdata.hascolumn(colname):
                    log.info(f"adding column {colname} to output table")
                    outdata.addcolumn(new_field_insert_point,  [colname])
            if pnotecolname and pnote:
                if not outdata.hascolumn(pnotecolname):
                    log.info(f"adding column {pnotecolname} to output table")
                    outdata.addcolumn(new_field_insert_point, [pnotecolname])
            if append_original_geodata_to_column:
                if not outdata.hascolumn(append_original_geodata_to_column):
                    log.info(f"adding column {append_original_geodata_to_column} to output table")
                    outdata.addcolumn(new_field_insert_point ,  [append_original_geodata_to_column]) 

            # Step through input file and process line by line
            # 1st line (header line) has already been read
            rowcount = 2
            while not indata.end(): 
                if (rowcount % 10) == 0: log.info(f"Processing row {rowcount}") 
                origdict = indata.next_row_as_dict() 
                outdict =  origdict.copy()
                edited = { k: False for k in outdict.keys() } # Edit status for each item on this row
                try: 
                    if rowcount < inc_first_data_line: # If this is a header line, just write
                        raise WriteRow
                    # If line has content in specified columns already, skip to WriteRow
                    for skipname in skip_if_content_columnnames: 
                        val = str(origdict.get(skipname,""))
                        if val and val.strip(): # Has some content
                            raise WriteRow 
                    matchrows = geodata.find_matches( origdict, rules, normalize_dict=regular_subs )
                    nmatch = len(matchrows)
                    if nmatch == 0:  
                        raise WriteRow
                    if nmatch > 1:  
                        log.debug(f"Found multiple matches for inputrow {rowcount}: {matchrows}. Check geodata source file. Skipping row")
                        raise WriteRow
                    # OK, so we have exactly one match
                    originaldata = [] # Kept to store original data from cells that may be replaced (for later reporting in the output)
                    mrow = matchrows[0] # index of the single matching row
                    match = geodata.get_result_dict(mrow, activeops) # match as a colname: val dictionary
                    for colname,val in match.items():
#                        print(colname,val )
                        if my2str(val).strip().lower() == knownd_keep: 
                            continue # Overrule marker in known_data
                        # If column name is not in outdata, it is not an active output field name and can be ignored
                        if colname.lower() not in outdata.lowercolnames: continue        
                        # Copy original data to a field in the output file (not copying the output cell data into itself
                        if append_original_geodata_to_column and (colname != append_original_geodata_to_column):  
                            oval = origdict.get(colname,"")  
                            if oval: originaldata.append(oval)
                        # OK to here
                        oper = geodata.get_output_action_for_column(colname, outputops) 
                        if oper not in outputops:
                            continue # Skip column with actions that are not output operations
                        elif (oper == cmd_replace) or ( oper == cmd_fillempty and not outdict[colname] ):
                            outdict[colname] = val
                            edited[colname] = op_replaced
                        elif oper == cmd_append and val: # Append non-empty values only
                            outdict[colname] = joinstr( outdict.get(colname,"" ),  val ,  itemsep ) 
                            edited[colname] = op_appended
                    if append_original_geodata_to_column: # Append old data to designated cell
                            origstr = f"{original_geodata_header} {itemsep.join(originaldata)}" 
                            cn = append_original_geodata_to_column.lower()
                            outdict[cn] = joinstr(outdict.get(cn,"" ),  origstr ,  "") 
                            edited[cn] = True
                    # Add note by the program, if available
                    if pnotecolname and pnote:
                        cn = pnotecolname.lower()
                        outdict[cn] = joinstr(outdict.get(cn,"" ),  pnote ,  itemsep) 
                        edited[cn] = True
                    raise WriteRow
                except WriteRow:
                    outdata.itersetrow(outdict,  edited)
                    next(outdata) # Move to next line in outdata
                    rowcount += 1
            log.info(f"Saving output file {outfn}")
            if not _SUPPRESS_FILE_CREATION_FOR_TESTING:
                outdata.close()
        except (jkError,  FileNotFoundError,  ValueError) as msg:
            log.critical(msg)
            sys.exit() 

    # Ask for any input before closing window
    # Now handled to a OS script wrapper (.bat on Windows)
